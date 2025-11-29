import os
import openpyxl
from PIL import Image
import io

def parse_data_type_from_checklist(checklist_file):
    r"""
    掃描工作表 H 欄 (由 H4 起) 取得資料型態。
    回傳格式: [{"row": <int>, "type": <str>} , ...]
    """
    wb = openpyxl.load_workbook(checklist_file)
    ws = wb.active
    data_type_location = []
    row = 4
    while True:
        h_cell = ws[f'H{row}']
        e_cell = ws[f'E{row}']
        if not h_cell.value and not e_cell.value:
            break
        if h_cell.value:
            data_type_location.append({"row": row, "type": str(h_cell.value).strip()})
        row += 1
    wb.close()
    return data_type_location

def parse_text_from_checklist(checklist_file, data_type_location):
    r"""
    依 data_type_location 取出對應列 E 欄文字 (僅 type 以 text: 開頭)。
    輸出檔案: input data/text/<checklist>_listed_text.txt
    回傳: {"file": <path>, "count": <int>, "lines": <list[str]>}
    """
    wb = openpyxl.load_workbook(checklist_file)
    ws = wb.active
    checklist_name = os.path.splitext(os.path.basename(checklist_file))[0]
    output_dir = os.path.join("input data", "text")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{checklist_name}_listed_text.txt")
    lines = []
    for item in data_type_location:
        if item["type"].startswith("text:"):
            row = item["row"]
            text = ws[f'E{row}'].value
            lang = item["type"].split(":", 1)[1].strip() if ":" in item["type"] else "eng"
            if text:
                lines.append(f"{text}, {lang}")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✓ 文字已輸出: {output_path} ({len(lines)} 筆)")
    wb.close()
    return {"file": output_path, "count": len(lines), "lines": lines}

def parse_image_from_checklist(checklist_file, data_type_location):
    r"""
    依 data_type_location 中 type == 'img' 的列匯出對應的內嵌圖片。
    目前以插入順序與 img 標記順序對應；若插圖數量多於標記或少於標記會提示。
    輸出檔案: input data/image/<checklist>_listed_img_XX.png
    回傳: {"files": <list[path]>, "requested": <int>, "exported": <int>, "embedded": <int>}
    """
    wb = openpyxl.load_workbook(checklist_file)
    ws = wb.active
    checklist_name = os.path.splitext(os.path.basename(checklist_file))[0]
    output_dir = os.path.join("input data", "image")
    os.makedirs(output_dir, exist_ok=True)
    img_rows = [item["row"] for item in data_type_location if item["type"] == "img"]
    images = getattr(ws, '_images', [])
    exported_files = []
    for idx, row in enumerate(img_rows):
        if idx >= len(images):
            print(f"⚠️  標記第 {row} 行無對應內嵌圖片 (標記多於圖片)")
            continue
        img = images[idx]
        try:
            img_data = img._data()
            pil_image = Image.open(io.BytesIO(img_data))
            out_path = os.path.join(output_dir, f"{checklist_name}_listed_img_{idx+1:02d}.png")
            pil_image.save(out_path, "PNG")
            print(f"✓ 影像已輸出: {out_path} ({pil_image.size[0]}x{pil_image.size[1]})")
            exported_files.append(out_path)
        except Exception as e:
            print(f"❌ 第 {row} 行圖片匯出失敗: {e}")
    if len(images) > len(img_rows):
        print(f"ℹ️  注意: 內嵌圖片共 {len(images)} 張，img 標記僅 {len(img_rows)} 行 (有未標記圖片)")
    wb.close()
    print(f"共輸出 {len(exported_files)} / {len(img_rows)} 張 (內嵌總數 {len(images)})")
    return {"files": exported_files, "requested": len(img_rows), "exported": len(exported_files), "embedded": len(images)}

def main(checklist_file):
    if not os.path.exists(checklist_file):
        raise FileNotFoundError(f"Checklist 檔案不存在: {checklist_file}")
    data_type_location = parse_data_type_from_checklist(checklist_file)
    text_info = parse_text_from_checklist(checklist_file, data_type_location)
    image_info = parse_image_from_checklist(checklist_file, data_type_location)
    summary = {
        "data_types": len(data_type_location),
        "text_count": text_info["count"],
        "image_requested": image_info["requested"],
        "image_exported": image_info["exported"],
        "image_embedded": image_info["embedded"],
        "text_file": text_info["file"],
    }
    print("\n✅ 摘要:")
    for k, v in summary.items():
        print(f"  - {k}: {v}")
    return summary

if __name__ == "__main__":
    CHECKLIST_FILE = r"input data\checklist\Checklist.xlsx"
    main(CHECKLIST_FILE)
    